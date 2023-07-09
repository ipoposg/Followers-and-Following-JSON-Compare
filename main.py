import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

def compare_json_files(followers_file, following_file):
    # Load the JSON data from the followers file
    with open(followers_file) as f:
        followers_data = json.load(f)

    # Load the JSON data from the following file
    with open(following_file) as f:
        following_data = json.load(f)

    # Extract the 'value' field from followers_data
    followers_values = [entry['string_list_data'][0]['value'] for entry in followers_data if entry.get('string_list_data')]

    # Extract the 'value' field from following_data
    following_values = [entry['string_list_data'][0]['value'] for entry in following_data['relationships_following']]

    # Find the values in following_values that are not in followers_values
    unique_values = [value for value in following_values if value not in followers_values]

    # Generate the output file name with current date and time
    current_datetime = datetime.now().strftime("%Y%m%d%H%M%S")
    output_file_name = f"output_{current_datetime}.xlsx"

    # Create an Excel workbook
    workbook = Workbook()
    sheet = workbook.active

    # Add headers
    headers = ['Value', 'REMOVED?']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f"{col_letter}1"] = header

    # Add unique values to the Excel sheet
    for row_num, value in enumerate(unique_values, 2):
        sheet[f"A{row_num}"] = value
        sheet[f"B{row_num}"] = ""  # Checkbox default value

    # Adjust column widths
    for col_num in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_num)
        sheet.column_dimensions[col_letter].width = 15

    # Save the Excel workbook
    workbook.save(output_file_name)

    print(f"Unique values saved to {output_file_name}")

    # Delete the followers and following files
    import os
    os.remove(followers_file)
    os.remove(following_file)
    print(f"{followers_file} and {following_file} deleted.")

# Replace the file names with the actual file paths
followers_file = "followers_1.json"
following_file = "following.json"

compare_json_files(followers_file, following_file)